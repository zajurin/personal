from bs4 import BeautifulSoup
import requests
import re
import phonenumbers
import urllib.request

import openpyxl
from openpyxl  import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from googletrans import Translator
import pycountry
from listOfCountries_and_codes import GetCountryCode

# #CREATING URL FOR GOOGLE SEARCHING

# my_Header
headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36'}


whichCountry = "US" # this variable MUST TO BE DINAMIC FROM Excel

companyName = 'North Star Charter School Inc' # this variable MUST TO BE DINAMIC FROM Excel
companyAddress = 'headquarters' # this variable MUST TO BE DINAMIC FROM Excel


#CREATING URL FOR GOOGLE SEARCHING
text = ("phone number of {0} {1}".format(companyName, companyAddress))# my_CopanyName & my_CompanyAddress

myCurrentURL = 'https://google.com/search?q=' + text

#ACCESING TO URL's INFO
response = requests.get(myCurrentURL, headers=headers)
soup = BeautifulSoup(response.text, 'html.parser')

search_classqrShPb_in_h2 = soup.find_all('h2', attrs={'class': 'qrShPb kno-ecr-pt PZPZlf mfMhoc hNKfZe'})# Level 1 - my_h2_className
search_CompanyName = soup.find_all('span', text = re.compile(companyName), attrs = {'class' : 'LC20lb DKV0Md'})# Level 2 - my_span_className

for single_h2 in search_classqrShPb_in_h2:
   first_span = single_h2.find('span').text	
   res_first_span = re.sub(r'[^\w\s]', '', first_span)
   res_companyName = re.sub(r'[^\w\s]', '', companyName)# my_CopanyName
  
   if res_first_span.lower() == res_companyName.lower():
      print("both are the same")
      find_country_in_the_box = soup.find('span', attrs={'class': 'LrzXr'})# my_span_className_in_BOX
      
      # Getting address from google box in this case is in spanish
      find_country_in_the_box_GET_TEXT = find_country_in_the_box.text
      # print(find_country_in_the_box_GET_TEXT)

      # Address might be in another language. So this must to be translated to english because pycountry works 
      # only with name of countries in english
      translator = Translator()

      # With "en" in << dest='en' >> we specified the country. 'en' stands for English language
      translate_text = translator.translate(find_country_in_the_box_GET_TEXT, dest='en')
      
      # Address translated to english and text is stored in a variable
      full_address_in_string = translate_text.text
      print(full_address_in_string)
      
      # In some cases pycountry recognize more than one country in the address. So we need to tell our program
      # Which one is the right one to be used in this program.
      # Fist of all, we need to create an array to store each country found it in the text box.
      # And then we'll take the last one because according to the format followed by google,
      # country name should be at the end of one address. At least inside of its boxes
      tiny_list_of_countries = []

      for single_country in pycountry.countries:

         if single_country.name in full_address_in_string:
            country_found = single_country.name
            # print(country_found)
            tiny_list_of_countries.append(country_found)

      if len(tiny_list_of_countries) > 1:         
         # print(tiny_list_of_countries[-1])
         chooseExcel_File = "C:\\Users\\rosenberg\\Desktop\\withPython\\PracticeBS4\\p1\\src\\myPhonenumbers.xlsx"

         # What's the name of the sheet where you are going to work?
         choose_SHEET_Of_Your_Excel_File = 'Hoja1'

         # Read an existing workbook
         wb = load_workbook(chooseExcel_File, data_only=True)
         sh = wb[choose_SHEET_Of_Your_Excel_File]

         rowM = sh.max_row

         # tell me in which Column (letter) and row (number) is the country name you're looking for
         concatenater_Column_With_Number_For_COUNTRY = 'E' + '9'# my_concatenater_Column_With_Number_For_COUNTRY

         # tell me the value inside of this cell and store it in whichCountry
         whichCountry = sh[concatenater_Column_With_Number_For_COUNTRY].value # my_WhichCountry

         # the [-1] section is because there is more than 1 country name in the text box of google
         # So we need to specify which one is the one needed. 
         # In this case is last one in the tiny_list_of_countries so we use [-1]
         if tiny_list_of_countries[-1].lower() == whichCountry.lower():# my_WhichCountry
            listOfCountries = pycountry.countries
            nameOfCountry = listOfCountries.get(name=whichCountry)# my_WhichCountry
            your_Country_Code = nameOfCountry.alpha_2
            
            # PhoneNumber library works with strings. So you need to turn your data into a string
            # after this process, you need to store it in a variable where PhoneNumber is going to
            # use the Matcher method
            # my_span_className_for_PhoneNumber
            new_SPAN_That_Needs_To_Be_STRING = str(soup.find('span', attrs={'class': 'LrzXr zdqRlf kno-fv'}))

            for match in phonenumbers.PhoneNumberMatcher(new_SPAN_That_Needs_To_Be_STRING, your_Country_Code):
               global only_Number
               only_Number = phonenumbers.format_number(match.number, phonenumbers.PhoneNumberFormat.E164)
               # Where are you gonna type Company phoneNumber
               number_of_ROW = '9'
               sh['F' + number_of_ROW] = only_Number # my_cell_to_type_PhoneNumber
               wb.save(filename = chooseExcel_File)


      elif len(tiny_list_of_countries) == 1:
         if tiny_list_of_countries.lower() == whichCountry.lower():# my_WhichCountry
            listOfCountries = pycountry.countries
            nameOfCountry = listOfCountries.get(name=whichCountry)# my_WhichCountry
            your_Country_Code = nameOfCountry.alpha_2

            new_SPAN_That_Needs_To_Be_STRING = str(soup.find('span', attrs={'class': 'LrzXr zdqRlf kno-fv'}))

            for match in phonenumbers.PhoneNumberMatcher(new_SPAN_That_Needs_To_Be_STRING, your_Country_Code):
               only_Number = phonenumbers.format_number(match.number, phonenumbers.PhoneNumberFormat.E164)
               # Where are you gonna type Company phoneNumber
               number_of_ROW = '9'
               sh['F' + number_of_ROW] = only_Number # my_cell_to_type_PhoneNumber
               wb.save(filename = chooseExcel_File)
         
         # print(single_country.name)
       
      else:
         print("nothing was found")  
              
   else:
      print("Are different")


