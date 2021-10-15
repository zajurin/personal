from bs4 import BeautifulSoup
import requests
import re
import phonenumbers
import urllib.request

import openpyxl
from openpyxl  import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

from googletrans import Translator
import pycountry
from listOfCountries_and_codes import GetCountryCode

class Another_2_constructor:
  def __init__(
    self, 
    my_Header, 
    my_WhichCountry, 
    my_CompanyName, 
    my_CompanyAddress, 
    my_h2_className,
    my_span_className,
    my_span_className_in_BOX,
    my_choose_SHEET_Of_Your_Excel_File,
    my_wb,
    my_sh,
    my_rowM,
    my_span_className_for_PhoneNumber,
    my_string_Of_number_Of_Row_With_Name_Of_Companies,
    my_chooseExcel_File):

    self.my_Header = my_Header
    self.my_WhichCountry = my_WhichCountry 
    self.my_CompanyName = my_CompanyName 
    self.my_CompanyAddress = my_CompanyAddress 
    self.my_h2_className = my_h2_className
    self.my_span_className = my_span_className
    self.my_span_className_in_BOX = my_span_className_in_BOX
    self.my_choose_SHEET_Of_Your_Excel_File = my_choose_SHEET_Of_Your_Excel_File
    self.my_wb = my_wb
    self.my_sh = my_sh
    self.my_rowM = my_rowM
    self.my_span_className_for_PhoneNumber = my_span_className_for_PhoneNumber
    self.my_string_Of_number_Of_Row_With_Name_Of_Companies = my_string_Of_number_Of_Row_With_Name_Of_Companies
    self.my_chooseExcel_File = my_chooseExcel_File

    # #CREATING URL FOR GOOGLE SEARCHING

    # my_Header
    # headers = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36'}


    # whichCountry = self.my_WhichCountry # this variable MUST TO BE DINAMIC FROM Excel

    # companyName = self.my_CompanyName # this variable MUST TO BE DINAMIC FROM Excel
    # companyAddress = self.my_CompanyAddress # this variable MUST TO BE DINAMIC FROM Excel


    #CREATING URL FOR GOOGLE SEARCHING
    text = ("phone number of {0} {1}".format(self.my_CompanyName, self.my_CompanyAddress))# my_CompanyName & my_CompanyAddress

    myCurrentURL = 'https://google.com/search?q=' + text

    #ACCESING TO URL's INFO
    response = requests.get(myCurrentURL, headers=self.my_Header)
    soup = BeautifulSoup(response.text, 'html.parser')

    # search_classqrShPb_in_h2 = soup.find_all('h2', attrs={'class': 'qrShPb kno-ecr-pt PZPZlf mfMhoc hNKfZe'})# Level 1 - my_h2_className
    search_classqrShPb_in_h2 = soup.find_all('h2', attrs={'class': self.my_h2_className})# Level 1 - my_h2_className

    # search_CompanyName = soup.find_all('span', text = re.compile(my_CompanyName), attrs = {'class' : 'LC20lb DKV0Md'})# Level 2 - my_span_className
    search_CompanyName = soup.find_all('span', text = re.compile(self.my_CompanyName), attrs = {'class' : self.my_span_className})# Level 2 - my_span_className

    for single_h2 in search_classqrShPb_in_h2:
       first_span = single_h2.find('span').text 
       res_first_span = re.sub(r'[^\w\s]', '', first_span)
       res_companyName = re.sub(r'[^\w\s]', '', self.my_CompanyName)# my_CompanyName
      
       # if res_first_span.lower() == res_companyName.lower():


       if res_first_span in res_companyName:
         global nameOfCompany_is_there
         nameOfCompany_is_there = True



       if nameOfCompany_is_there:
          print("both are the same")
          # find_country_in_the_box = soup.find('span', attrs={'class': 'LrzXr'})# my_span_className_in_BOX
          find_country_in_the_box = soup.find('span', attrs={'class': self.my_span_className_in_BOX})# my_span_className_in_BOX
          
          # Getting address from google box in this case is in spanish
          find_country_in_the_box_GET_TEXT = find_country_in_the_box.text
          # print(find_country_in_the_box_GET_TEXT)

          # Address might be in another language. So this must to be translated to english because pycountry works 
          # only with name of countries in english
          translator = Translator()

          # With "en" in << dest='en' >> we specified the country. 'en' stands for English language
          translate_text = translator.translate(find_country_in_the_box_GET_TEXT, dest='en')
          
          # Address translated to english and text is stored in a variable
          full_address_in_string = translate_text.text
          print(full_address_in_string)
          
          # In some cases pycountry recognize more than one country in the address. So we need to tell our program
          # Which one is the right one to be used in this program.
          # Fist of all, we need to create an array to store each country found it in the text box.
          # And then we'll take the last one because according to the format followed by google,
          # country name should be at the end of one address. At least inside of its boxes
          tiny_list_of_countries = []

          for single_country in pycountry.countries:

             if single_country.name in full_address_in_string:
                country_found = single_country.name
                # print(country_found)
                tiny_list_of_countries.append(country_found)

          if len(tiny_list_of_countries) > 1:         
             # print(tiny_list_of_countries[-1])
             # chooseExcel_File = "C:\\Users\\rosenberg\\Desktop\\withPython\\PracticeBS4\\p1\\src\\myPhonenumbers.xlsx"

             # What's the name of the sheet where you are going to work?
             # choose_SHEET_Of_Your_Excel_File = 'Hoja1' # my_choose_SHEET_Of_Your_Excel_File
             choose_SHEET_Of_Your_Excel_File = self.my_choose_SHEET_Of_Your_Excel_File # self.my_choose_SHEET_Of_Your_Excel_File

             # Read an existing workbook
             # wb = load_workbook(chooseExcel_File, data_only=True)# my_wb
             # sh = wb[choose_SHEET_Of_Your_Excel_File]# my_sh

             # rowM = sh.max_row

             # tell me the value inside of this cell and store it in whichCountry
             # whichCountry = sh[concatenater_Column_With_Number_For_COUNTRY].value # my_WhichCountry

             # the [-1] section is because there is more than 1 country name in the text box of google
             # So we need to specify which one is the one needed. 
             # In this case is last one in the tiny_list_of_countries so we use [-1]
             if tiny_list_of_countries[-1].lower() == self.my_WhichCountry.lower():# my_WhichCountry
                listOfCountries = pycountry.countries
                nameOfCountry = listOfCountries.get(name=self.my_WhichCountry)# my_WhichCountry
                your_Country_Code = nameOfCountry.alpha_2
                
                # PhoneNumber library works with strings. So you need to turn your data into a string
                # after this process, you need to store it in a variable where PhoneNumber is going to
                # use the Matcher method
                # new_SPAN_That_Needs_To_Be_STRING = str(soup.find('span', attrs={'class': 'LrzXr zdqRlf kno-fv'}))# my_span_className_for_PhoneNumber
                new_SPAN_That_Needs_To_Be_STRING = str(soup.find('span', attrs={'class': self.my_span_className_for_PhoneNumber}))# my_span_className_for_PhoneNumber

                for match in phonenumbers.PhoneNumberMatcher(new_SPAN_That_Needs_To_Be_STRING, your_Country_Code):
                   global only_Number
                   only_Number = phonenumbers.format_number(match.number, phonenumbers.PhoneNumberFormat.E164)
                   # Where are you gonna type Company phoneNumber
                   number_of_ROW = self.my_string_Of_number_Of_Row_With_Name_Of_Companies
                   self.my_sh['F' + number_of_ROW] = only_Number # my_cell_to_type_PhoneNumber
                   self.my_wb.save(filename = self.my_chooseExcel_File)


             elif len(tiny_list_of_countries) == 1:
               global tinyListVariable
               tinyListVariable = tiny_list_of_countries[-1].lower()

               global dynamicVariable
               dynamicVariable = self.my_WhichCountry[-1].lower()
              


             if tiny_list_of_countries[-1].lower() == self.my_WhichCountry.lower():# my_WhichCountry
                listOfCountries = pycountry.countries
                nameOfCountry = listOfCountries.get(name=self.my_WhichCountry)# my_WhichCountry
                your_Country_Code = nameOfCountry.alpha_2

                new_SPAN_That_Needs_To_Be_STRING = str(soup.find('span', attrs={'class': 'LrzXr zdqRlf kno-fv'}))

                for match in phonenumbers.PhoneNumberMatcher(new_SPAN_That_Needs_To_Be_STRING, your_Country_Code):
                   only_Number = phonenumbers.format_number(match.number, phonenumbers.PhoneNumberFormat.E164)
                   # Where are you gonna type Company phoneNumber
                   number_of_ROW = self.my_string_Of_number_Of_Row_With_Name_Of_Companies
                   self.my_sh['F' + number_of_ROW] = only_Number # my_cell_to_type_PhoneNumber
                   self.my_wb.save(filename = self.my_chooseExcel_File)
             
             # print(single_country.name)
           
          else:
             print("nothing was found")  
                  
       else:
          print("Are different")
          
