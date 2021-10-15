import openpyxl
from openpyxl  import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation

wb2 = load_workbook('.xlsx')
print(wb2)



# # Create a data-validation object with list validation
# dv = DataValidation(type="list", formula1='"Dog,Cat,Bat"', allow_blank=True)

# # Optionally set a custom error message
# dv.error ='Your entry is not in the list'
# dv.errorTitle = 'Invalid Entry'

# # Optionally set a custom prompt message
# dv.prompt = 'Please select from the list'
# dv.promptTitle = 'List Selection'

# # Add the data-validation object to the worksheet
# ws.add_data_validation(dv)

# # Create some cells, and add them to the data-validation object
# c1 = ws["A1"]
# c1.value = "Dog"
# dv.add(c1)
# c2 = ws["A2"]
# c2.value = "An invalid value"
# dv.add(c2)

# # Or, apply the validation to a range of cells
# dv.add('B1:B1048576') # This is the same as for the whole of column B

# # Check with a cell is in the validator
# "B4" in dv